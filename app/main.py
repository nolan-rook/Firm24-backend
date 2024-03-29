import os
from fastapi import FastAPI, HTTPException, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from orquesta_sdk import Orquesta, OrquestaClientOptions
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# Initialize FastAPI app
app = FastAPI()

# Set up CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Initialize Orquesta client
def init_orquesta_client():
    api_key = os.getenv("ORQUESTA_API_KEY")
    options = OrquestaClientOptions(api_key=api_key, environment="production")
    return Orquesta(options)

client = init_orquesta_client()

current_question_indices = {}

def load_questions_from_sheet(sheet_path):
    questions_with_options = []
    workbook = load_workbook(sheet_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=0, values_only=True):  # Assuming row 1 has the headers
        question = row[0]  # Question is in the first column
        quick_reply_options = []
        if row[1]:  # Check if there are quick reply options in the second column
            # Split options on semicolon and strip whitespace
            quick_reply_options = [option.strip() for option in row[1].split(';')]
        
        if question:
            # We will not use question_index or condition as they are not present in the Excel file
            questions_with_options.append((None, question, quick_reply_options, None))
    print(questions_with_options)
    return questions_with_options

# Load the questions and options when the app starts
questions_with_options = load_questions_from_sheet("data/Firm24_lijst.xlsx")

@app.post("/question/")
async def question(request: Request):
    data = await request.json()
    user_id = data.get("user_id")
    question_index = data.get("question_index")
    previous_question = data.get("previous_question")
    previous_answer = data.get("previous_answer")

    print(f"Received question_index={question_index}, previous_answer='{previous_answer}'")

    if user_id in current_question_indices:
        # Resend the original question with rephrasing after clarification
        stored_question_index = current_question_indices[user_id]
        _, question, quick_reply_options, _ = questions_with_options[stored_question_index - 1]
        
        # Proceed to rephrase the question before resending
        rephrased_question = client.deployments.invoke(
            key="Firm24_vragenlijst",
            context={"environments": []},
            inputs={"question": question, "previous": ""}
        ).choices[0].message.content
        
        del current_question_indices[user_id]  # Clear stored index after use

        return {"rephrased_question": rephrased_question, "quick_reply_options": quick_reply_options}

    evaluation_deployment = client.deployments.invoke(
        key="Firm24-evaluate-user-input",
        context={"language": ["Dutch"]},
        inputs={"previous_question": previous_question, "previous_answer": previous_answer}
    )
    evaluation_result = evaluation_deployment.choices[0].message.content

    if evaluation_result == "Nee":
        current_question_indices[user_id] = question_index

        handle_answer_deployment = client.deployments.invoke(
            key="Firm24-handle-clarification",
            inputs={"previous_answer": previous_answer, "previous_question": previous_question}
        )
        handle_answer = handle_answer_deployment.choices[0].message.content

        return {"rephrased_question": handle_answer, "quick_reply_options": []}

    # Fetch and rephrase the next question if the answer is "Ja" or it's a new question
    if question_index is None or question_index < 1 or question_index > len(questions_with_options):
        raise HTTPException(status_code=400, detail="Invalid question index")

    if question_index <= len(questions_with_options):
        _, next_question, quick_reply_options, _ = questions_with_options[question_index - 1]
        
        previous_context = f"Vraag: {previous_question}\nAntwoord: {previous_answer}" if previous_question and previous_answer else ""
        
        rephrased_question = client.deployments.invoke(
            key="Firm24_vragenlijst",
            context={"environments": []},
            inputs={"question": next_question, "previous": previous_context}
        ).choices[0].message.content

        print(f"Sending to frontend: rephrased_question='{rephrased_question}', quick_reply_options={quick_reply_options}")

        return {"rephrased_question": rephrased_question, "quick_reply_options": quick_reply_options}
    else:
        raise HTTPException(status_code=400, detail="No suitable question found")

def is_condition_met(condition, previous_answer, combined_questions_with_options):
    condition_question_index, valid_answers = condition.split('=')
    # Split the valid answers by comma to support multiple valid answers
    valid_answers_list = valid_answers.split(',')
    # No need to adjust for zero-based indexing if your question indices already start at 1
    return previous_answer in valid_answers_list

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)